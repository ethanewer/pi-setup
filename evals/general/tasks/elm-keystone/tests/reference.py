#!/usr/bin/env python3
"""elm-keystone reference generator + comparator (verifier side).

Commands:
  gen <input.csv> <outdir>   : produce the canonical expected artifacts
  cmp <agent_dir> <exp_dir>  : deep-compare an agent output dir to the expected;
                              exit 0 iff identical (prints COMPARE-OK).

This is the reference behaviour the agent's /app/solve.py must reproduce. It is
independent of the solution implementation. All rules mirror instruction.md.
"""
import csv
import json
import os
import sys
from collections import defaultdict


def parse_requests(path):
    """Return only the valid rows; malformed/duplicate rows are dropped."""
    seen = set()
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            try:
                rid = (raw.get("route_id") or "").strip()
                zone = (raw.get("zone") or "").strip()
                slot = int((raw.get("slot") or "").strip())
                units = int((raw.get("units") or "").strip())
                cost = int((raw.get("cost") or "").strip())
            except (ValueError, TypeError):
                continue                       # missing/malformed column -> invalid
            if not rid or not zone:
                continue
            if not (0 <= slot <= 47) or units <= 0 or cost <= 0:
                continue
            if rid in seen:
                continue                       # duplicate route_id -> skip repeat
            seen.add(rid)
            rows.append({"route_id": rid, "zone": zone, "slot": slot,
                         "units": units, "cost": cost, "value": units * cost})
    return rows


def capacity(zone):
    return 15 + len(zone)                       # deterministic from zone name


def best_subset(zone, items):
    """0/1 knapsack: maximize sum(units*cost) s.t. sum(units) <= capacity(zone).

    Tie-break: among subsets with max value, choose the one whose sorted tuple
    of route_ids is lexicographically smallest.
    """
    cap = capacity(zone)
    n = len(items)
    best_val = None
    best_key = None
    for mask in range(1 << n):
        w = 0
        v = 0
        chosen = []
        for i in range(n):
            if (mask >> i) & 1:
                w += items[i]["units"]
                v += items[i]["value"]
                chosen.append(items[i]["route_id"])
        if w > cap:
            continue
        key = tuple(sorted(chosen))
        if best_val is None or v > best_val or (v == best_val and key < best_key):
            best_val = v
            best_key = key
    return best_val, set(best_key)


def compute(rows):
    by_zone = defaultdict(list)
    for r in rows:
        by_zone[r["zone"]].append(r)
    decision = {}                # route_id -> RUN/HOLD
    zone_report = {}             # zone -> {"opt", "planned", "served"}
    chosen_routes = []           # route dicts with decision RUN
    for zone in sorted(by_zone):
        items = by_zone[zone]
        opt, chosen = best_subset(zone, items)
        planned = 0
        served = 0
        for r in items:
            dec = "RUN" if r["route_id"] in chosen else "HOLD"
            decision[r["route_id"]] = dec
            if dec == "RUN":
                planned += r["units"]
                served += 1
                chosen_routes.append(r)
        zone_report[zone] = {"opt": opt, "planned": planned, "served": served}
    objective = sum(z["opt"] for z in zone_report.values())
    return decision, zone_report, chosen_routes, objective


def artefact_xlsx(rows, decision, chosen_routes, objective, out_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws["A1"] = "DEPLOYMENT_SCHEDULE"
    for col, hdr in zip("ABCDEF", ["route_id", "zone", "slot", "units",
                                   "cost_unit", "value"]):
        ws.cell(2, "ABCDEF".index(col) + 1, hdr)
    plan = sorted(chosen_routes, key=lambda r: (r["zone"], r["slot"], r["route_id"]))
    r = 3
    for p in plan:
        ws.cell(r, 1, p["route_id"])
        ws.cell(r, 2, p["zone"])
        ws.cell(r, 3, p["slot"])
        ws.cell(r, 4, p["units"])
        ws.cell(r, 5, p["cost"])
        ws.cell(r, 6, p["value"])
        r += 1
    label_row = len(plan) + 4                     # data rows 3..(len+2), blank, then label
    ws.cell(label_row, 1, "OPTIMAL_OBJECTIVE")
    ws.cell(label_row, 2, objective)

    fl = wb.create_sheet("Flags")
    fl["A1"], fl["B1"] = "route_id", "flag"
    ordered = sorted(rows, key=lambda r: (r["zone"], r["slot"], r["route_id"]))
    for i, rr in enumerate(ordered, start=2):
        fl.cell(i, 1, rr["route_id"])
        fl.cell(i, 2, decision[rr["route_id"]])

    wb.save(out_path)


def gen(input_path, outdir):
    rows = parse_requests(input_path)
    decision, zone_report, chosen_routes, objective = compute(rows)

    os.makedirs(outdir, exist_ok=True)

    with open(os.path.join(outdir, "plan_records.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["route_id", "zone", "slot", "units", "cost", "value", "decision"])
        for r in rows:
            w.writerow([r["route_id"], r["zone"], r["slot"], r["units"],
                        r["cost"], r["value"], decision[r["route_id"]]])

    with open(os.path.join(outdir, "decisions.txt"), "w") as f:
        f.write("objective=%d\n" % objective)
        for rid in sorted(decision):
            f.write("%s=%s\n" % (rid, decision[rid]))

    with open(os.path.join(outdir, "objective.txt"), "w") as f:
        f.write("objective=%d\n" % objective)

    with open(os.path.join(outdir, "final_report.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["zone", "planned_units", "opt_value", "served"])
        for zname in sorted(zone_report):
            z = zone_report[zname]
            w.writerow([zname, z["planned"], z["opt"], z["served"]])

    tdir = os.path.join(outdir, "transformed")
    os.makedirs(tdir, exist_ok=True)
    grp = defaultdict(list)
    for r in rows:
        grp[r["zone"]].append(r)
    for zone in sorted(grp):
        lines = sorted(grp[zone], key=lambda r: (r["slot"], r["route_id"]))
        with open(os.path.join(tdir, zone.lower() + ".csv"), "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["route_id", "slot", "units", "value", "decision"])
            for r in lines:
                w.writerow([r["route_id"], r["slot"], r["units"],
                            r["value"], decision[r["route_id"]]])

    plan = sorted(chosen_routes, key=lambda r: (r["zone"], r["slot"], r["route_id"]))
    answer = {
        "schema_version": 1,
        "objective": objective,
        "plan": [{
            "route_id": r["route_id"], "zone": r["zone"], "slot": r["slot"],
            "units": r["units"], "cost_unit": r["cost"], "decision": "RUN"
        } for r in plan],
    }
    with open(os.path.join(outdir, "answer.json"), "w") as f:
        json.dump(answer, f, sort_keys=True)

    artefact_xlsx(rows, decision, chosen_routes, objective, os.path.join(outdir, "schedule.xlsx"))
    return objective


def _files_flat(dirpath):
    """Recursively list regular files under dirpath (relative names)."""
    out = []
    for here, _dirs, files in os.walk(dirpath):
        for fn in files:
            rel = os.path.relpath(os.path.join(here, fn), dirpath)
            out.append(rel)
    return sorted(out)


def _read_csv(p):
    with open(p, newline="") as f:
        return list(csv.reader(f))


def cmp(agent_dir, exp_dir):
    fails = []
    exp = _files_flat(exp_dir)
    agent = _files_flat(agent_dir)
    if agent != exp:
        fails.append("file set %s != %s" % (agent, exp))
    for name in exp:
        apath = os.path.join(agent_dir, name)
        if not os.path.exists(apath):
            fails.append("missing %s" % name)
            continue
        epath = os.path.join(exp_dir, name)
        if name == "answer.json":
            if json.load(open(apath)) != json.load(open(epath)):
                fails.append("answer.json mismatch (schema-exact)")
        elif name == "schedule.xlsx":
            from openpyxl import load_workbook
            try:
                aw = load_workbook(apath, data_only=True)
                ew = load_workbook(epath, data_only=True)
                if aw.sheetnames != ew.sheetnames:
                    fails.append("xlsx sheetnames %s != %s" % (aw.sheetnames, ew.sheetnames))
                for sn in ew.sheetnames:
                    wsa, wse = aw[sn], ew[sn]
                    for row in wse.iter_rows():
                        for c in row:
                            if c.value is None:
                                continue
                            av = wsa.cell(c.row, c.column).value
                            if av != c.value:
                                fails.append("xlsx %s!%s%d: %r != %r"
                                             % (sn, c.coordinate[:1], c.row, av, c.value))
            except Exception as ex:
                fails.append("xlsx compare error: %s" % ex)
        elif name.endswith(".csv"):
            if _read_csv(apath) != _read_csv(epath):
                fails.append("CSV %s differs (schema/order/cell)" % name)
        else:  # text: exact lines, trailing whitespace normalized, blank lines ignored
            def lines(p):
                return [ln.rstrip() for ln in open(p).read().splitlines() if ln.strip()]
            if lines(apath) != lines(epath):
                fails.append("text %s differs" % name)
    if fails:
        for f in fails:
            print("COMPARE-FAIL: " + f)
        return False
    print("COMPARE-OK")
    return True


if __name__ == "__main__":
    mode = sys.argv[1]
    if mode == "gen":
        gen(sys.argv[2], sys.argv[3])
    elif mode == "cmp":
        sys.exit(0 if cmp(sys.argv[2], sys.argv[3]) else 1)