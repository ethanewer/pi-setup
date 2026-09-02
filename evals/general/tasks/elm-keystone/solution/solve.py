#!/usr/bin/env python3
"""Keystone deployment-plan solver.

Reads a fleet-request CSV and emits a byte-exact, schema-exact deployment plan
artifacts pack: plan_records.csv, decisions.txt, objective.txt,
final_report.csv, transformed/<zone>.csv, answer.json and schedule.xlsx.

Usage:
    python3 solve.py --input requests.csv [--output /app]

Rules (see /app/instruction.md in the task):
  * Records are row-invalid (dropped) when a required column is missing or
    non-integer, route_id/zone is blank, slot is outside 0..47, units or cost
    is <= 0, or the route_id was already seen earlier in the file.
  * Each zone is solved as a 0/1 knapsack maximizing sum(units*cost) subject to
    sum(units) <= capacity(zone), where capacity(zone) = 15 + len(zone).
  * Tie-break: among subsets with maximum value, choose the one whose sorted
    tuple of route_ids is lexicographically smallest (=> deterministic).
"""
import argparse
import csv
import json
import os
import sys
from collections import defaultdict


def parse_requests(path):
    rows = []
    seen = set()
    with open(path, newline="", encoding="utf-8") as f:
        for raw in csv.DictReader(f):
            try:
                rid = (raw.get("route_id") or "").strip()
                zone = (raw.get("zone") or "").strip()
                slot = int((raw.get("slot") or "").strip())
                units = int((raw.get("units") or "").strip())
                cost = int((raw.get("cost") or "").strip())
            except (ValueError, TypeError):
                continue
            if not rid or not zone:
                continue
            if not (0 <= slot <= 47) or units <= 0 or cost <= 0:
                continue
            if rid in seen:
                continue
            seen.add(rid)
            rows.append({"route_id": rid, "zone": zone, "slot": slot,
                         "units": units, "cost": cost, "value": units * cost})
    return rows


def capacity(zone):
    return 15 + len(zone)


def best_subset(zone, items):
    cap = capacity(zone)
    n = len(items)
    best_val = None
    best_key = None
    for mask in range(1 << n):
        w = 0
        v = 0
        chosen = []
        for i in range(n):
            if (mask >> i) & 1:
                w += items[i]["units"]
                v += items[i]["value"]
                chosen.append(items[i]["route_id"])
        if w > cap:
            continue
        key = tuple(sorted(chosen))
        if best_val is None or v > best_val or (v == best_val and key < best_key):
            best_val = v
            best_key = key
    return best_val, set(best_key)


def build(rows):
    by_zone = defaultdict(list)
    for r in rows:
        by_zone[r["zone"]].append(r)
    decision = {}
    zone_report = {}
    chosen_routes = []
    for zone in sorted(by_zone):
        items = by_zone[zone]
        opt, chosen = best_subset(zone, items)
        planned = 0
        served = 0
        for r in items:
            dec = "RUN" if r["route_id"] in chosen else "HOLD"
            decision[r["route_id"]] = dec
            if dec == "RUN":
                planned += r["units"]
                served += 1
                chosen_routes.append(r)
        zone_report[zone] = {"opt": opt, "planned": planned, "served": served}
    objective = sum(z["opt"] for z in zone_report.values())
    return decision, zone_report, chosen_routes, objective


def write_outputs(rows, decision, zone_report, chosen_routes, objective, outdir):
    os.makedirs(outdir, exist_ok=True)

    with open(os.path.join(outdir, "plan_records.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["route_id", "zone", "slot", "units", "cost", "value", "decision"])
        for r in rows:
            w.writerow([r["route_id"], r["zone"], r["slot"], r["units"],
                        r["cost"], r["value"], decision[r["route_id"]]])

    with open(os.path.join(outdir, "decisions.txt"), "w") as f:
        f.write("objective=%d\n" % objective)
        for rid in sorted(decision):
            f.write("%s=%s\n" % (rid, decision[rid]))

    with open(os.path.join(outdir, "objective.txt"), "w") as f:
        f.write("objective=%d\n" % objective)

    with open(os.path.join(outdir, "final_report.csv"), "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["zone", "planned_units", "opt_value", "served"])
        for zname in sorted(zone_report):
            z = zone_report[zname]
            w.writerow([zname, z["planned"], z["opt"], z["served"]])

    from collections import defaultdict
    grp = defaultdict(list)
    for r in rows:
        grp[r["zone"]].append(r)
    tdir = os.path.join(outdir, "transformed")
    os.makedirs(tdir, exist_ok=True)
    for zone in sorted(grp):
        lines = sorted(grp[zone], key=lambda r: (r["slot"], r["route_id"]))
        with open(os.path.join(tdir, zone.lower() + ".csv"), "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["route_id", "slot", "units", "value", "decision"])
            for r in lines:
                w.writerow([r["route_id"], r["slot"], r["units"],
                            r["value"], decision[r["route_id"]]])

    plan = sorted(chosen_routes, key=lambda r: (r["zone"], r["slot"], r["route_id"]))
    answer = {
        "schema_version": 1,
        "objective": objective,
        "plan": [{
            "route_id": r["route_id"], "zone": r["zone"], "slot": r["slot"],
            "units": r["units"], "cost_unit": r["cost"], "decision": "RUN"
        } for r in plan],
    }
    with open(os.path.join(outdir, "answer.json"), "w") as f:
        json.dump(answer, f, sort_keys=True)

    write_xlsx(rows, decision, chosen_routes, objective,
               os.path.join(outdir, "schedule.xlsx"))


def write_xlsx(rows, decision, chosen_routes, objective, out_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws["A1"] = "DEPLOYMENT_SCHEDULE"
    for col, hdr in zip("ABCDEF", ["route_id", "zone", "slot", "units",
                                   "cost_unit", "value"]):
        ws.cell(2, "ABCDEF".index(col) + 1, hdr)
    plan = sorted(chosen_routes, key=lambda r: (r["zone"], r["slot"], r["route_id"]))
    r = 3
    for p in plan:
        ws.cell(r, 1, p["route_id"])
        ws.cell(r, 2, p["zone"])
        ws.cell(r, 3, p["slot"])
        ws.cell(r, 4, p["units"])
        ws.cell(r, 5, p["cost"])
        ws.cell(r, 6, p["value"])
        r += 1
    label_row = len(plan) + 4
    ws.cell(label_row, 1, "OPTIMAL_OBJECTIVE")
    ws.cell(label_row, 2, objective)

    fl = wb.create_sheet("Flags")
    fl["A1"], fl["B1"] = "route_id", "flag"
    ordered = sorted(rows, key=lambda r: (r["zone"], r["slot"], r["route_id"]))
    for i, rr in enumerate(ordered, start=2):
        fl.cell(i, 1, rr["route_id"])
        fl.cell(i, 2, decision[rr["route_id"]])
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=os.environ.get("KS_INPUT", "/opt/keystone/requests.csv"))
    ap.add_argument("--output", default=os.environ.get("KS_OUTPUT", "/app"))
    args = ap.parse_args()
    rows = parse_requests(args.input)
    decision, zone_report, chosen_routes, objective = build(rows)
    write_outputs(rows, decision, zone_report, chosen_routes, objective, args.output)
    print("objective=%d plans=%d" % (objective, len(chosen_routes)))


if __name__ == "__main__":
    main()