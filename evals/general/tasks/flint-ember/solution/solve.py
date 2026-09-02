#!/usr/bin/env python3
"""ZephyrGrid standardized batch-pipeline reference implementation.

Reads <workdir>/input/ (projects.csv + config.json) and writes the exact
required artifacts into <workdir>:

  plans.jsonl            plan records (id, batch, shape)
  decision.txt           invest/defer flags + maximised NPV block
  result.csv             exact-final-column CSV (string-typed)
  output/results.json    salt / iterations / deterministic file hex index
  answer.json            optimal integer objective (max benefit), no newline
  ledger.xlsx            values written via openpyxl at exact cell addresses

Usage:  python3 solve.py [workdir]     (default workdir = /app)
"""
import csv
import hashlib
import json
import math
import os
import sys

from openpyxl import Workbook

ROOT = "/app"


def prefers_not_investing(lhs, rhs, n):
    """True iff between two equal-benefit masks, lhs is the one preferred by
    the documented tie rule: prefer NOT investing lower-index (earlier-row)
    projects, i.e. at the lowest index where the masks differ lhs has 0."""
    for i in range(n):
        li = (lhs >> i) & 1
        ri = (rhs >> i) & 1
        if li != ri:
            return li == 0
    return False


def knapsack_opt(projects, budget):
    """Maximise total benefit with total cost <= budget (0/1 knapsack).

    Enumerates all subsets (tiny instances guaranteed); on an exact benefit
    tie, prefer NOT investing lower-index (earlier-row) projects.
    """
    n = len(projects)
    best_b = -1
    best_mask = -1
    for mask in range(1 << n):
        cost = 0
        benefit = 0
        for i in range(n):
            if (mask >> i) & 1:
                cost += int(projects[i]['cost'])
                benefit += int(projects[i]['benefit'])
        if cost <= budget and (
            benefit > best_b
            or (benefit == best_b and prefers_not_investing(mask, best_mask, n))
        ):
            best_b = benefit
            best_mask = mask
    chosen = [bool((best_mask >> i) & 1) for i in range(n)]
    return chosen, best_b


def npv_of(p):
    revenue = float(p['revenue'])
    life = int(p['life'])
    rate = float(p['rate'])
    cost = float(p['cost'])
    total = sum(revenue / (1.0 + rate) ** t for t in range(1, life + 1))
    return total - cost


def hex_index(rel, content, salt, iterations):
    """Deterministic per-file hex index: sha256 of (salt|path|bytes) then
    re-hashed `iterations` times."""
    payload = (salt + "|" + rel).encode("utf-8") + content
    h = hashlib.sha256(payload).hexdigest()
    for _ in range(max(0, iterations - 1)):
        h = hashlib.sha256(h.encode("ascii")).hexdigest()
    return h


def main(workdir):
    os.makedirs(os.path.join(workdir, "input"), exist_ok=True)
    out_json_dir = os.path.join(workdir, "output")
    os.makedirs(out_json_dir, exist_ok=True)

    with open(os.path.join(workdir, "input", "config.json")) as fh:
        config = json.load(fh)
    salt = config["salt"]
    iterations = int(config["iterations"])
    budget = int(config["budget"])

    with open(os.path.join(workdir, "input", "projects.csv"), newline="") as fh:
        projects = list(csv.DictReader(fh))

    chosen, best_b = knapsack_opt(projects, budget)
    npvs = []
    for p in projects:
        npvs.append(npv_of(p))
    total_npv = sum(v for i, v in enumerate(npvs) if chosen[i])

    # ---- 1. plans.jsonl ----------------------------------------------------
    with open(os.path.join(workdir, "plans.jsonl"), "w") as fh:
        for p in projects:
            shape = {
                "capacity": float(p["capacity"]),
                "life": int(p["life"]),
                "revenue": float(p["revenue"]),
            }
            rec = {
                "id": p["id"],
                "batch": p["batch"],
                "shape": shape,
            }
            fh.write(json.dumps(rec, separators=(",", ":")) + "\n")

    # ---- 2. decision.txt ---------------------------------------------------
    lines = []
    for i, p in enumerate(projects):
        invest = "yes" if chosen[i] else "no"
        defer = "no" if chosen[i] else "yes"
        lines.append("project=%s invest=%s defer=%s" % (p["id"], invest, defer))
    lines.append("npv_total=%.2f" % total_npv)
    with open(os.path.join(workdir, "decision.txt"), "w") as fh:
        fh.write("\n".join(lines) + "\n")

    # ---- 3. result.csv -----------------------------------------------------
    with open(os.path.join(workdir, "result.csv"), "w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["id", "invest", "defer", "benefit", "cost"])
        for i, p in enumerate(projects):
            invest = "yes" if chosen[i] else "no"
            defer = "no" if chosen[i] else "yes"
            writer.writerow([p["id"], invest, defer, p["benefit"], p["cost"]])

    # ---- 4. answer.json (optimal integer objective, no trailing newline) ---
    with open(os.path.join(workdir, "answer.json"), "w") as fh:
        fh.write(str(best_b))

    # ---- 5. ledger.xlsx (exact cell addresses via openpyxl) ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "ledger"
    headers = ["id", "invest", "defer", "benefit", "cost", "npv"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)
    for i, p in enumerate(projects):
        row = i + 2
        invest = "yes" if chosen[i] else "no"
        defer = "no" if chosen[i] else "yes"
        ws.cell(row=row, column=1, value=p["id"])
        ws.cell(row=row, column=2, value=invest)
        ws.cell(row=row, column=3, value=defer)
        ws.cell(row=row, column=4, value=int(p["benefit"]))
        ws.cell(row=row, column=5, value=int(p["cost"]))
        ws.cell(row=row, column=6, value=round(npvs[i], 2))
    sr = len(projects) + 3
    ws.cell(row=sr, column=1, value="total_benefit")
    ws.cell(row=sr, column=2, value=best_b)
    ws.cell(row=sr + 1, column=1, value="total_npv")
    ws.cell(row=sr + 1, column=2, value=round(total_npv, 2))
    wb.save(os.path.join(workdir, "ledger.xlsx"))

    # ---- 6. output/results.json (salt, iterations, sorted file hex map) ----
    rels = [
        "answer.json",
        "decision.txt",
        "input/projects.csv",
        "plans.jsonl",
        "result.csv",
    ]
    files = []
    for rel in sorted(rels):
        with open(os.path.join(workdir, rel), "rb") as fh:
            data = fh.read()
        files.append({"path": rel, "hex": hex_index(rel, data, salt, iterations)})
    results = {"salt": salt, "iterations": iterations, "files": files}
    with open(os.path.join(out_json_dir, "results.json"), "w") as fh:
        json.dump(results, fh, indent=2)
        fh.write("\n")


if __name__ == "__main__":
    workdir = sys.argv[1] if len(sys.argv) > 1 else "/app"
    main(workdir)